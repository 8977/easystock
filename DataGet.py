import tkinter as tk
from tkinter import scrolledtext,END
import requests as re
import openpyxl
import datetime

#建立窗口主体
window=tk.Tk()

#确定窗口名称
window.title('指数爬取')

#设定窗口大小
win_w=420
win_h=450
window.geometry("{}x{}".format(win_w, win_h))

#获取窗口大小


# 定义一个标签和输入组合
class LabelEntry():
    #定义类接口变量
    def __init__(self,Text,LabelEntry_x,LabelEntry_y,L_width,E_width):
        self.Text=Text
        self.LabelEntry_x=LabelEntry_x
        self.LabelEntry_y=LabelEntry_y
        self.L_width=L_width
        self.E_width=E_width
        
    #定义内部变量
    Item_Label=tk.Label()
    Item_split=tk.Label()
    Item_Entry=tk.Entry()

    #展示对应变量
    def Display_LE(self):
        self.Item_Label=tk.Label(window,text=self.Text,font=('Times New Roman', 14),justify='left',width=self.L_width, height=1,anchor='e')
        self.Item_split=tk.Label(window,text=':',font=('Times New Roman', 14),width=1, height=1,anchor='n')
        self.Item_Entry=tk.Entry(window,show=None,width=self.E_width)
        self.Item_Label.place(x=self.LabelEntry_x,y=self.LabelEntry_y,anchor='nw')
        self.Item_split.place(x=self.LabelEntry_x+self.L_width*11,y=self.LabelEntry_y,anchor='nw')
        self.Item_Entry.place(x=self.LabelEntry_x+self.L_width*13,y=self.LabelEntry_y,anchor='nw')


    #获取输入文本
    def GetEntry(self):
        t=self.Item_Entry.get()
        #清理Item_Entry
        self.Item_Entry.delete(0, END)
        #将Item_Entry接收到的文本插入Item_Entry
        self.Item_Entry.insert(10, t)



def GetText():
    with open("datasave.txt", "r",encoding='UTF-8') as f:
        data = f.readlines()
        for i in range(len(data)):
            data[i]=data[i].split('\n')[0]
    return data

#按钮函数
def undo():
    for i in range(6):
        LE_list[i].Item_Entry.delete(0,'end')
        data=GetText()
        LE_list[i].Item_Entry.insert(40,data[i])
    History_text.insert('end','===== 已撤销 =====\n')
    History_text.see(END) #聚焦历史记录末尾
    
def enter():
    for i in range(6):
        t=LE_list[i].Item_Entry.get()
        if data[i]!=t:
            History_text.insert('end',Label_list[i]+' changed:\nOLD:'+data[i]+'\nNEW:'+t+'\n')
            data[i]=t
    History_text.see(END)
            
def textsave():
    with open("datasave.txt", "w",encoding='UTF-8') as f:
        for i in range(len(data)):
            f.write(data[i]+'\n')
    History_text.see('end')
            
def dataget():
    for i in range(3):
        t=datasave(data[2*i],data[2*i+1])
        History_text.insert('end',Label_list[2*i+1]+':'+t+'\n')
    History_text.see(END)

#爬取程序
def datasave(lujing,xzcode):

    def datacatch(xzcode,starttime,endtime):
        url='http://quotes.money.163.com/service/chddata.html?code='+xzcode+'&start='+starttime+'&end='+endtime+'&fields=TCLOSE;HIGH;LOW;TOPEN;LCLOSE;CHG;PCHG;VOTURNOVER;VATURNOVER'

        try:
            r=re.get(url,timeout=30)
            r.raise_for_status()
            r.encoding=r.apparent_encoding
            return r.text
        except:
            return '爬取失败'

    def workbookload(lujing):
        try:
            WorkBook=openpyxl.load_workbook(lujing)
            return WorkBook
        except:
            return '文件读取失败'
        
    xzcode=xzcode
    
    WorkBook=workbookload(lujing)
    
    if WorkBook !='文件读取失败':
        WorkSheet=WorkBook.worksheets[0]
        #读取起始时间
        StartTime_0=WorkSheet.cell(2,1).value
        StartTime_1=StartTime_0+datetime.timedelta(days=1)
        EndTime=datetime.datetime.now()

        starttime=StartTime_1.strftime('%Y%m%d')
        endtime=EndTime.strftime('%Y%m%d')
        #爬取文件
        rtext=datacatch(xzcode,starttime,endtime)
        
        if rtext != '爬取失败':
            rtextlist=rtext.split('\n')
            t=len(rtextlist)-1
            for g in range(t):
                if g>0 and g<t:
                    rtextcells=rtextlist[len(rtextlist)-g-1].split(',')
                   
                    WorkSheet.insert_rows(2)

                    for i in range(len(rtextcells)):
                        if i==0:
                            s=datetime.datetime.strptime(rtextcells[i],'%Y-%m-%d')
                        elif i>2:
                            s=eval(rtextcells[i])
                        else:
                            s=rtextcells[i]
                        if i==0:
                            WorkSheet.cell(2,i+1).value=s
                            WorkSheet.cell(2,i+1).number_format='YYYY/M/D'
                        else:
                            WorkSheet.cell(2,i+1).value=s
            WorkBook.save(lujing)
            return 'SaveOk'
        else:
            return '爬取失败'
    else:
        return '文件读取失败'


#左对齐线
LeftLine=10

#item存入列表并展示，entry应display之后
Label_list=['沪深300','codeHS300','上证指数','codeSZ','深成指数','codeSC']
data=GetText()
LE_list=[]
for i in range(6):
    t=LabelEntry(Label_list[i],LeftLine,i*24+14,9,40)
    LE_list.append(t)
    #t.Item_Entry.bind("<Return>", t.GetEntry)
    
#记录LabelEntry的Bottom
BottomLine=i*24+38

#显示历史信息
History_list=[]
History_text=scrolledtext.ScrolledText(window,height=15,width=55)
History_text.insert(tk.END,'历史信息:\n')

#按钮
Button_undo=tk.Button(window,text='撤销',width=8,command=undo)
Button_enter=tk.Button(window,text='确认',width=8,command=enter)
Button_dataget=tk.Button(window,text='爬取',width=8,command=dataget)
Button_textsave=tk.Button(window,text='确定修改',width=8,command=textsave)
Button_list=[Button_undo,Button_enter,Button_textsave,Button_dataget]
Button_num=len(Button_list)
#界面布局
#LE
for i in range(6):
    LE_list[i].Display_LE()
    LE_list[i].Item_Entry.insert(40,data[i]) #insrt在布局之后
#历史信息
History_text.place(x=LeftLine,y=BottomLine+10,anchor='nw')
#按钮
for i in range(Button_num):
    Button_list[Button_num-i-1].place(x=win_w-66*i-80,y=win_h-40,anchor='nw')

while __name__=='__main__':   
    window.mainloop()
